import pdfplumber
import pandas as pd
import os
import re
import calendar
import sys 
from datetime import datetime
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter

# --- 경로 설정 ---
def get_base_path():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    else:
        return os.path.dirname(os.path.abspath(__file__))

base_path = get_base_path()
folder_path = os.path.join(base_path, "관내출장복명서전용")
output_dir = base_path

if not os.path.exists(folder_path):
    os.makedirs(folder_path, exist_ok=True)
    print(f"📁 '관내출장복명서전용' 폴더를 생성했습니다. PDF를 넣고 다시 실행하세요.")

# 직위 매핑
POSITION_MAP = {
    "이상일": "본부장",
    "박기수": "단장", "김경열": "선임", "손정훈": "인턴", "김유정": "인턴",
    "황수경": "전임", "황선주": "전임", "이진리": "전임", "박보배": "전임",
    "임건우": "전임",  "김연미": "전임", "오희웅": "전임", "장진우": "전임"
} 

def calculate_travel_fee(duration_text):
    try:
        times = re.findall(r'\d{1,2}:\d{2}', duration_text)
        if len(times) >= 2:
            fmt = "%H:%M"
            start_t = datetime.strptime(times[0], fmt)
            end_t = datetime.strptime(times[1], fmt)
            diff = end_t - start_t
            hours = diff.total_seconds() / 3600
            if hours >= 4: return 20000
            elif 0 < hours < 4: return 10000
    except: pass
    return 0

def get_info_from_table(pdf):
    extracted_info = {"period": None, "purpose": None, "reg_no": None}
    try:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                for row in table:
                    clean_row = [cell.strip().replace(" ", "") if cell else "" for cell in row]
                    if not extracted_info["period"]:
                        for i, text in enumerate(clean_row):
                            if "출장기간" in text or "일시" in text:
                                if (i + 1) < len(row) and row[i+1]:
                                    extracted_info["period"] = row[i+1].replace("\n", " ").strip()
                    if not extracted_info["purpose"]:
                        for i, text in enumerate(clean_row):
                            if "출장목적" in text:
                                if (i + 1) < len(row) and row[i+1]:
                                    extracted_info["purpose"] = re.sub(r'^[:\s\-\.○□■▷▶*]+', '', row[i+1].replace("\n", " ")).strip()
                    if not extracted_info["reg_no"]:
                        for i, text in enumerate(clean_row):
                            if "등록번호" in text:
                                if (i + 1) < len(row) and row[i+1]:
                                    extracted_info["reg_no"] = str(row[i+1]).replace(" ", "").replace("\n", "")[:14]
    except: pass
    return extracted_info

def extract_smart(full_text, file_name, table_info=None):
    file_name_fixed = os.path.splitext(file_name)[0]
    base_data = {
        "파일명": file_name_fixed, "출장자": "", "직위": "",
        "출발지": "우리원", "도착지": "우리원",
        "출장기간": "", "출장지": "", "출장목적": "", "출장내용": "",
        "출장비": 0, "등록번호": ""
    }
    
    match = re.search(r'\([월화수목금토일]\)\s*,\s*(.*?)\)[^(]*$', file_name_fixed)
    if match: base_data["출장지"] = match.group(1).strip()
    
    if table_info:
        base_data["출장기간"] = table_info.get("period") or ""
        base_data["출장목적"] = table_info.get("purpose") or ""
        base_data["등록번호"] = table_info.get("reg_no") or ""

    lines = [line.strip() for line in full_text.split('\n') if line.strip()]
    names_list = []
    
    for line in lines:
        clean_line = line.replace(" ", "")  # 띄어쓰기 우선 모두 제거
        if "성명" in clean_line:
            # "성명"을 기준으로 문자열을 쪼개고, 그 뒤에 오는 텍스트만 타겟으로 삼음
            target_text = clean_line.split("성명")[-1] 
            
            for known_name in POSITION_MAP.keys():
                if known_name in target_text:
                    names_list.append(known_name)
                    
    unique_names = list(dict.fromkeys(names_list))
    base_data["출장비"] = calculate_travel_fee(base_data["출장기간"])
    
    results = []
    if not unique_names:
        results.append(base_data)
    else:
        for name in unique_names:
            new_row = base_data.copy()
            new_row["출장자"] = name
            new_row["직위"] = POSITION_MAP.get(name, "전임")
            results.append(new_row)
    return results

def save_monthly_excel(df, year, month):
    output_filename = f"{month}월 관내여비지급내역서({month}월).xlsx"
    output_xlsx = os.path.join(output_dir, output_filename)
    
    df = df.sort_values(by=["출장자", "출장기간"])
    df["출장비"] = pd.to_numeric(df["출장비"], errors='coerce').fillna(0)
    
    _, last_day = calendar.monthrange(year, month)
    period_text = f"□ 기간 : {year}년 {month}월 1일 ~ {year}년 {month}월 {last_day}일까지"

    final_rows = []
    total_sum = df["출장비"].sum()
    # 합계 행
    final_rows.append({"순번": "합 계", "파일명": "", "출장기간": "", "출발지": "", "경유지": "", "도착지": "", "직위": "", "출장자": "", "출장비": total_sum, "영수인": "", "비고": "", "등록번호": ""})

    for name, group in df.groupby("출장자"):
        # 소계 행
        final_rows.append({"순번": "소계", "파일명": "", "출장기간": "", "출발지": "", "경유지": "", "도착지": name, "직위": "", "출장자": "", "출장비": group["출장비"].sum(), "영수인": "", "비고": "", "등록번호": ""})
        for idx, (_, r) in enumerate(group.iterrows(), 1):
            item = r.to_dict()
            item["순번"] = idx
            item["경유지"] = item.pop("출장지", "")
            item["비고"] = item.pop("출장목적", "")
            item["영수인"] = ""
            final_rows.append(item)

    excel_df = pd.DataFrame(final_rows)
    cols = ["순번", "파일명", "출장기간", "출발지", "경유지", "도착지", "직위", "출장자", "출장비", "영수인", "비고", "등록번호"]
    excel_df = excel_df[cols]

    try:
        with pd.ExcelWriter(output_xlsx, engine='openpyxl') as writer:
            excel_df.to_excel(writer, index=False, sheet_name='Sheet1', startrow=5, startcol=1, header=False)
            ws = writer.sheets['Sheet1']
            
            # 스타일 설정
            font_header = Font(name='HY헤드라인M', size=11, bold=True)
            border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
            
            # 상단 제목 영역
            ws.merge_cells('D1:M1'); ws['D1'] = f"{month}월 관내여비지급내역"; ws['D1'].font = Font(name='HY헤드라인M', size=20, bold=True); ws['D1'].alignment = Alignment(horizontal='center')
            ws.merge_cells('D2:F2'); ws['D2'] = "□ 부서 : AI산업진흥단"; ws['D2'].font = Font(name='돋움', size=11); ws['D2'].alignment = Alignment(horizontal='center')
            ws.merge_cells('D3:I3'); ws['D3'] = period_text; ws['D3'].font = Font(name='돋움', size=11); ws['D3'].alignment = Alignment(horizontal='center')

            # 헤더(4~5행)
            headers_4 = {2:"순번", 3:"파일명", 4:"출장기간", 5:"출발지", 6:"경유지", 7:"도착지", 8:"금 액", 11:"영수인", 12:"비고", 13:"등록번호"}
            headers_5 = {8:"직위", 9:"성명", 10:"금액"}
            for c, v in headers_4.items(): ws.cell(4, c).value = v
            for c, v in headers_5.items(): ws.cell(5, c).value = v
            for r in [4, 5]:
                for c in range(2, 14):
                    cell = ws.cell(r, c)
                    cell.font = font_header; cell.border = border; cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            
            for c in [2,3,4,5,6,7,11,12,13]: ws.merge_cells(start_row=4, start_column=c, end_row=5, end_column=c)
            ws.merge_cells(start_row=4, start_column=8, end_row=4, end_column=10)

            # 데이터 행 서식
            for r_idx in range(6, ws.max_row + 1):
                ws.row_dimensions[r_idx].height = 35
                val = ws.cell(r_idx, 2).value
                is_total = val in ["합 계", "소계"]
                if is_total:
                    ws.merge_cells(start_row=r_idx, start_column=2, end_row=r_idx, end_column=6)
                    ws.merge_cells(start_row=r_idx, start_column=7, end_row=r_idx, end_column=9)

                for c_idx in range(2, 14):
                    cell = ws.cell(r_idx, c_idx)
                    cell.border = border; cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    if is_total: cell.font = Font(bold=True)
                    if c_idx == 10: cell.number_format = '#,##0'

            # 너비 조절
            widths = [7, 40, 30, 7, 25, 7, 7, 7, 10, 6, 30, 20]
            for i, w in enumerate(widths, 2): ws.column_dimensions[get_column_letter(i)].width = w

        print(f"🎉 생성 완료: {output_filename}")
    except Exception as e:
        print(f"❌ 저장 실패: {e}")

def run_total_extraction(folder):
    pdf_files = [f for f in os.listdir(folder) if f.lower().endswith('.pdf')]
    if not pdf_files:
        print("🔍 '관내출장복명서' 폴더에 파일이 없습니다.")
        return

    print(f"🚀 총 {len(pdf_files)}개의 파일을 처음부터 다시 읽습니다...")
    all_results = []
    for file_name in pdf_files:
        try:
            with pdfplumber.open(os.path.join(folder, file_name)) as pdf:
                t_info = get_info_from_table(pdf)
                txt = "".join([(p.extract_text() or "") for p in pdf.pages])
                all_results.extend(extract_smart(txt, file_name, t_info))
                print(f"✅ 읽기 성공: {file_name}")
        except Exception as e: print(f"❌ 오류({file_name}): {e}")

    if all_results:
        m_df = pd.DataFrame(all_results)
        # 날짜 추출 및 연/월 분류 (공백 허용 정규식 및 공백 제거 로직 적용)
        m_df['parsed_date'] = pd.to_datetime(
            m_df['출장기간']
            .str.extract(r'(\d{2,4}[\.-]\s*\d{1,2}[\.-]\s*\d{1,2})')[0]
            .str.replace(r'\s+', '', regex=True)
            .str.replace('.', '-'),
            errors='coerce'
        ).fillna(datetime.now())
        
        m_df['year'], m_df['month'] = m_df['parsed_date'].dt.year, m_df['parsed_date'].dt.month
        
        for (y, m), group in m_df.groupby(['year', 'month']):
            save_monthly_excel(group, int(y), int(m))

if __name__ == "__main__":
    run_total_extraction(folder_path)
    input("\n모든 작업이 완료되었습니다. 엔터를 눌러 종료하세요.")