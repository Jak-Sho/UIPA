import pdfplumber
import pandas as pd
import os
import re
import calendar
import sys 
from datetime import datetime
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter

# --- 경로 설정 구간 ---
def get_base_path():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    else:
        return os.path.dirname(os.path.abspath(__file__))

base_path = get_base_path()
folder_path = os.path.join(base_path, "PDF")
log_file_path = os.path.join(base_path, "처리완료목록.txt")
output_dir = base_path

if not os.path.exists(folder_path):
    os.makedirs(folder_path, exist_ok=True)
    print(f"📁 알림: 실행 위치에 'PDF' 폴더를 생성했습니다.\n   경로: {folder_path}\n   이 폴더에 PDF 파일을 넣고 다시 실행해주세요.")

# ----------------------------------------------------

POSITION_MAP = {
    "이상일": "본부장",
    "박기수": "단장", "김경열": "선임", "손정훈": "인턴", "김유정": "인턴",
    "황수경": "전임", "황선주": "전임", "이진리": "전임", "박보배": "전임",
    "임건우": "전임",  "김연미": "전임", "오희웅": "전임", "장진우": "전임"
} 

def calculate_travel_fee(duration_text):
    try:
        times = re.findall(r'\d{1,2}:\d{2}', duration_text)
        if len(times) >= 2:
            fmt = "%H:%M"
            start_t = datetime.strptime(times[0], fmt)
            end_t = datetime.strptime(times[1], fmt)
            if end_t < start_t: pass
            diff = end_t - start_t
            hours = diff.total_seconds() / 3600
            if hours >= 4: return 20000
            elif 0 < hours < 4: return 10000
    except Exception:
        pass
    return 0

def get_info_from_table(pdf):
    # 등록번호(reg_no) 항목 추가
    extracted_info = {"period": None, "purpose": None, "reg_no": None}
    try:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                for row in table:
                    clean_row = [cell.strip().replace(" ", "") if cell else "" for cell in row]
                    
                    if not extracted_info["period"]:
                        for i, text in enumerate(clean_row):
                            if "출장기간" in text or "일시" in text:
                                if (i + 1) < len(row) and row[i+1]:
                                    merged_text = row[i+1].replace("\n", " ").strip()
                                    if any(char.isdigit() for char in merged_text):
                                        extracted_info["period"] = merged_text
                                break
                    if not extracted_info["purpose"]:
                        for i, text in enumerate(clean_row):
                            if "출장목적" in text:
                                if (i + 1) < len(row) and row[i+1]:
                                    merged_purpose = row[i+1].replace("\n", " ").strip()
                                    merged_purpose = re.sub(r'^[:\s\-\.○□■▷▶*]+', '', merged_purpose).strip()
                                    extracted_info["purpose"] = merged_purpose
                                break
                    
                    # --- 등록번호 추출 로직 추가 (방법 A: 모든 공백 제거 후 14자) ---
                    if not extracted_info["reg_no"]:
                        for i, text in enumerate(clean_row):
                            if "등록번호" in text:
                                if (i + 1) < len(row) and row[i+1]:
                                    raw_reg = str(row[i+1]).strip()
                                    clean_reg = raw_reg.replace(" ", "").replace("\n", "")
                                    extracted_info["reg_no"] = clean_reg[:14]
                                break

            if all(extracted_info.values()):
                break
    except Exception as e:
        print(f"⚠️ 표 데이터 추출 중 오류: {e}")
    return extracted_info

def extract_smart(full_text, file_name, table_info=None):
    file_name_fixed = os.path.splitext(file_name)[0]
    base_data = {
        "파일명": file_name_fixed, "출장자": "", "직위": "",
        "출발지": "우리원", "도착지": "우리원",
        "출장기간": "", "출장지": "", "출장목적": "", "출장내용": "",
        "출장비": 0, "등록번호": "" # 필드 추가
    }

    try:
        # (월)~(일) 요일 패턴 뒤의 쉼표부터, 제일 마지막 닫는 괄호 사이의 모든 텍스트를 탐색
        match = re.search(r'\([월화수목금토일]\)\s*,\s*(.*?)\)[^(]*$', file_name_fixed)
        if match:
            place_extracted = match.group(1).strip()
            if place_extracted:
                base_data["출장지"] = place_extracted
    except Exception:
        pass

    if table_info:
        if table_info.get("period"): base_data["출장기간"] = table_info["period"]
        if table_info.get("purpose"): base_data["출장목적"] = table_info["purpose"]
        if table_info.get("reg_no"): base_data["등록번호"] = table_info["reg_no"] # 등록번호 반영

    lines = [line.strip() for line in full_text.split('\n') if line.strip()]
    names_list = []
    last_info_idx = 0

    has_date_format = False
    if base_data["출장기간"]:
        if re.search(r'\d{2,4}[\.-]\d{1,2}[\.-]\d{1,2}', base_data["출장기간"].replace(" ", "")):
            has_date_format = True

    if not has_date_format:
        for line in lines:
            clean_line = line.replace(" ", "")
            date_match = re.search(r'(\d{2,4}[\.-]\d{1,2}[\.-]\d{1,2})', clean_line)
            if date_match:
                found_date = date_match.group(1)
                if base_data["출장기간"]:
                    base_data["출장기간"] = found_date + " " + base_data["출장기간"]
                else:
                    base_data["출장기간"] = found_date
                break

    for i, line in enumerate(lines):
        if not base_data["출장기간"]:
            combined_match = re.search(r'일\s*시\s*/\s*장\s*소', line)
            if combined_match:
                raw_content = line[combined_match.end():].strip()
                content = re.sub(r'^[:\s]+', '', raw_content)
                if "/" in content:
                    parts = content.split("/")
                    base_data["출장기간"] = parts[0].strip()
                    if not base_data["출장지"]: base_data["출장지"] = "/".join(parts[1:]).strip()
                else:
                    base_data["출장기간"] = content
                last_info_idx = i
                continue

            time_match = re.search(r'일\s*시', line)
            if time_match:
                if "/" not in line[:time_match.end()+5]: 
                    content = line[time_match.end():].strip()
                    base_data["출장기간"] = re.sub(r'^[:\s]+', '', content)
                    last_info_idx = i

        place_match = re.search(r'장\s*소', line)
        if place_match and not base_data["출장지"]:
            content = line[place_match.end():].strip()
            if not re.search(r'\d{4}[\.-]\d{2}[\.-]\d{2}', content):
                base_data["출장지"] = re.sub(r'^[:\s]+', '', content)
                last_info_idx = i

        if not base_data["출장목적"]:
            purpose_match = re.search(r'출\s*장\s*목\s*적', line)
            if purpose_match:
                purpose = line[purpose_match.end():].strip()
                base_data["출장목적"] = re.sub(r'^[:\s\-\.○□■▷▶*]+', '', purpose).strip()

        name_keyword_match = re.search(r'성\s*명', line)
        if name_keyword_match:
            raw_names_content = line[name_keyword_match.end():].strip()
            raw_names_content = re.sub(r'^[:\s]+', '', raw_names_content)
            search_idx = i + 1
            while search_idx < len(lines):
                next_line = lines[search_idx]
                if "귀하" in next_line or re.search(r'(일\s*시|장\s*소|목\s*적|주\s*요\s*내\s*용)', next_line):
                    break
                raw_names_content += " " + next_line
                search_idx += 1
            if raw_names_content:
                for known_name in POSITION_MAP.keys():
                    spaced_pattern = r'\s*'.join(list(known_name))
                    if re.search(spaced_pattern, raw_names_content):
                        raw_names_content = re.sub(spaced_pattern, known_name, raw_names_content)
                tokens = re.split(r'[,/.\s]+', raw_names_content)
                for t in tokens:
                    c = t.strip()
                    if c and 2 <= len(c) <= 4 and "귀하" not in c: 
                        names_list.append(c)

    if base_data["출장기간"]: 
        base_data["출장비"] = calculate_travel_fee(base_data["출장기간"])
    
    target_pattern = r'(출\s*장\s*내\s*용|주\s*요\s*내\s*용|주\s*요\s*회\s*의\s*내\s*용|출\s*장\s*결\s*과|주\s*요\s*활\s*동\s*내\s*용)'
    summary_line = ""
    search_range = lines[last_info_idx:]
    for j, line in enumerate(search_range):
        kw_match = re.search(target_pattern, line)
        if kw_match:
            after_kw = line[kw_match.end():].strip()
            clean_after = re.sub(r'^[:\s\-\.○□■▷▶*]+', '', after_kw).strip()
            if clean_after and len(clean_after) > 1: summary_line = clean_after
            elif (j + 1) < len(search_range):
                next_line = search_range[j+1]
                if "위와 같이" not in next_line:
                    summary_line = re.sub(r'^[:\s\-\.○□■▷▶*]+', '', next_line).strip()
            if summary_line: break 
    base_data["출장내용"] = summary_line

    results = []
    unique_names = list(dict.fromkeys(names_list))
    if not unique_names: results.append(base_data)
    else:
        for name in unique_names:
            new_row = base_data.copy()
            new_row["출장자"] = name
            new_row["직위"] = POSITION_MAP.get(name, "전임")
            results.append(new_row)
    return results

def save_monthly_excel(df, year, month):
    output_filename = f"{month}월 관내여비지급내역서({month}월).xlsx"
    output_xlsx = os.path.join(output_dir, output_filename)

    if os.path.exists(output_xlsx):
        try:
            ex_df = pd.read_excel(output_xlsx)
            final_save_df = pd.concat([ex_df, df], ignore_index=True)
            print(f"ℹ️ [업데이트] {output_filename} 에 {len(df)}건 추가")
        except: final_save_df = df
    else:
        final_save_df = df
        print(f"ℹ️ [신규생성] {output_filename}")

    final_save_df = final_save_df.sort_values(by=["출장자", "출장기간"], ascending=[True, True])
    _, last_day = calendar.monthrange(year, month)
    period_text = f"□ 기간 : {year}년 {month}월 1일 ~ {year}년 {month}월 {last_day}일까지"

    final_rows = []
    total_fee_sum = final_save_df["출장비"].sum()
    grand_total_row = {
        "순번": "합 계", "파일명": "", "출장기간": "", "출발지": "", "경유지": "",
        "도착지": "", "직위": "", "출장자": "", "출장비": total_fee_sum, "영수인": "", "비고": "", "등록번호": ""
    }
    final_rows.append(grand_total_row)

    for name, group in final_save_df.groupby("출장자"):
        subtotal_row = {
            "순번": "소계", "파일명": "", "출장기간": "", "출발지": "", "경유지": "",           
            "도착지": f"{name}", "직위": "", "출장자": "", 
            "출장비": group["출장비"].sum(), "영수인": "", "비고": "", "등록번호": ""
        }
        final_rows.append(subtotal_row)
        for idx, (_, row_data) in enumerate(group.iterrows(), 1):
            item = row_data.to_dict()
            item["순번"] = idx
            item["경유지"] = item.pop("출장지", "")
            item["비고"] = item.pop("출장목적", "") 
            item["영수인"] = "" 
            final_rows.append(item)

    excel_df = pd.DataFrame(final_rows)
    # M열(등록번호)까지 컬럼 정의
    cols = ["순번", "파일명", "출장기간", "출발지", "경유지", "도착지", "직위", "출장자", "출장비", "영수인", "비고", "등록번호"]
    excel_df = excel_df[cols]

    try:
        with pd.ExcelWriter(output_xlsx, engine='openpyxl') as writer:
            excel_df.to_excel(writer, index=False, sheet_name='Sheet1', startrow=5, startcol=1, header=False)
            ws = writer.sheets['Sheet1']
            
            font_title = Font(name='HY헤드라인M', size=20, bold=True)
            font_dotum = Font(name='돋움', size=11, bold=False)
            font_header = Font(name='HY헤드라인M', size=11, bold=True)
            font_normal = Font(name='맑은 고딕', size=10, bold=False)
            font_date = Font(name='맑은 고딕', size=10, bold=False)
            font_bold = Font(name='맑은 고딕', size=11, bold=True)
            font_confirm = Font(name='HyhwpEQ', size=12, bold=False)
            font_sign = Font(name='HyhwpEQ', size=14, bold=True)
            very_thick = Side(style='thick', color="000000")
            border = Border(left=very_thick, right=very_thick, top=very_thick, bottom=very_thick)
            fill_gray = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

            # 제목 및 상단 (M열까지 확장)
            ws.merge_cells('D1:M1')
            ws['D1'].value = f"{month}월 관내여비지급내역"
            ws['D1'].font = font_title
            ws['D1'].alignment = Alignment(horizontal='center', vertical='center')
            
            ws.merge_cells('D2:F2')
            ws['D2'].value = "□ 부서 : AI산업진흥단"
            ws['D2'].font = font_dotum
            ws['D2'].alignment = Alignment(horizontal='center', vertical='center')
            
            ws.merge_cells('D3:I3')
            ws['D3'].value = period_text
            ws['D3'].font = font_dotum
            ws['D3'].alignment = Alignment(horizontal='center', vertical='center')

            ws.column_dimensions['A'].width = 5
            # M열(등록번호) 너비 20 추가
            widths = [7, 45, 16, 7, 28, 7, 7, 7, 10, 6, 40, 20]
            for i, w in enumerate(widths, 2):
                ws.column_dimensions[get_column_letter(i)].width = w

            # 헤더 정의
            headers_4 = {2:"순번", 3:"파일명", 4:"출장기간", 5:"출발지", 6:"경유지", 7:"도착지", 8:"금 액", 11:"영수인", 12:"비고", 13:"등록번호"}
            headers_5 = {8:"직위", 9:"성명", 10:"금액"}
            for c, v in headers_4.items(): ws.cell(4, c).value = v
            for c, v in headers_5.items(): ws.cell(5, c).value = v
            
            for r in [4, 5]:
                ws.row_dimensions[r].height = 16
                for c in range(2, 14): # M열까지
                    cell = ws.cell(r, c)
                    cell.font = font_header
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = border
                    cell.fill = fill_gray
            
            for c in [2,3,4,5,6,7,11,12,13]: ws.merge_cells(start_row=4, start_column=c, end_row=5, end_column=c)
            ws.merge_cells(start_row=4, start_column=8, end_row=4, end_column=10)

            for r_idx in range(6, ws.max_row + 1):
                ws.row_dimensions[r_idx].height = 35
                val = ws.cell(r_idx, 2).value
                
                if val in ["합 계", "소계"]:
                    ws.merge_cells(start_row=r_idx, start_column=2, end_row=r_idx, end_column=6)
                    ws.merge_cells(start_row=r_idx, start_column=7, end_row=r_idx, end_column=9)
                
                is_sub_total = (val in ["합 계", "소계"])
                
                for c_idx in range(2, 14): # M열까지
                    cell = ws.cell(r_idx, c_idx)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = border
                    if is_sub_total:
                        cell.fill = fill_gray
                        cell.font = font_bold
                    else:
                        cell.font = font_date if c_idx == 4 else font_normal
                    if c_idx == 10: cell.number_format = '#,##0'

            footer_r = ws.max_row + 1
            ws.row_dimensions[footer_r].height = 35
            ws.cell(footer_r, 2).value = "위와 같이 출장하였음을 확인함."
            ws.merge_cells(start_row=footer_r, start_column=2, end_row=footer_r, end_column=7)
            for c in range(2, 14):
                cell = ws.cell(footer_r, c)
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = font_confirm
            
            sign_r = footer_r + 2
            ws.merge_cells(start_row=sign_r, start_column=9, end_row=sign_r, end_column=13)
            sign_cell = ws.cell(sign_r, 9)
            sign_cell.value = "확인자  :  AI산업진흥단장 박기수     (인)"
            sign_cell.font = font_sign
            sign_cell.alignment = Alignment(horizontal='right', vertical='center')

        print(f"🎉 저장 완료: {output_filename}")
    except Exception as e:
        print(f"❌ 저장 실패 ({output_filename}): {e}")

def run_total_extraction(folder):
    if not os.path.exists(folder): return
    processed_files = set()
    if os.path.exists(log_file_path):
        with open(log_file_path, "r", encoding="utf-8") as f:
            processed_files = set(line.strip() for line in f)

    pdf_files = [f for f in os.listdir(folder) if f.lower().endswith('.pdf') and f not in processed_files]
    if not pdf_files:
        print("\n🔍 처리할 새로운 PDF 파일이 없습니다.")
        return

    print(f"🔍 {len(pdf_files)}개의 파일을 처리합니다...")
    all_results = []
    newly_processed = []

    for file_name in pdf_files:
        full_path = os.path.join(folder, file_name)
        try:
            with pdfplumber.open(full_path) as pdf:
                t_info = get_info_from_table(pdf)
                txt = ""
                for p in pdf.pages: txt += (p.extract_text() or "") + "\n"
                
                extracted_data = extract_smart(txt, file_name, t_info)
                all_results.extend(extracted_data)
                
                newly_processed.append(file_name)
                print(f"✅ 추출 성공: {file_name} [등록번호: {t_info.get('reg_no') or '없음'}]")
        except Exception as e:
            print(f"❌ 오류({file_name}): {e}")

    if all_results:
        master_df = pd.DataFrame(all_results)
        
        # 날짜 파싱 (UserWarning 방지를 위한 정돈된 로직)
        clean_dates = master_df['출장기간'].astype(str).str.replace(' ', '')
        extracted_dates = clean_dates.str.extract(r'(\d{2,4}[\.-]\d{1,2}[\.-]\d{1,2})')[0]
        
        master_df['parsed_date'] = pd.to_datetime(
            extracted_dates.str.replace('.', '-'), 
            errors='coerce'
        )
        
        mask = master_df['parsed_date'].dt.year < 100
        master_df.loc[mask, 'parsed_date'] = master_df.loc[mask, 'parsed_date'] + pd.DateOffset(years=2000)
        master_df['parsed_date'] = master_df['parsed_date'].fillna(datetime.now())
        
        master_df['year'] = master_df['parsed_date'].dt.year
        master_df['month'] = master_df['parsed_date'].dt.month

        for (year, month), group_df in master_df.groupby(['year', 'month']):
            print(f"\n📂 {year}년 {month}월 데이터 저장 시작...")
            save_monthly_excel(group_df, int(year), int(month))

        with open(log_file_path, "a", encoding="utf-8") as f:
            for name in newly_processed: f.write(name + "\n")

if __name__ == "__main__":
    from openpyxl.utils import get_column_letter # 추가 임포트
    run_total_extraction(folder_path)
    input("\n작업이 완료되었습니다. 종료하려면 엔터 키를 누르세요...")
